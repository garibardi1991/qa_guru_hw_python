import csv
import zipfile
from openpyxl import load_workbook
from pypdf import PdfReader
from script_os import ZIP_DIR


def test_pdf_file():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zip_file:
        with zip_file.open("PDF_Converter_Pro_Quick_Reference_Guide.RU.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[2]
            text = page.extract_text()
            assert '3Почему  популярны  файлы  PDF?\n' in text


def test_xlsx_file():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zip_file:
        with zip_file.open("test_file.xlsx") as excel_file:
            wb = load_workbook(filename=excel_file)
            sheet = wb.active
            cell_value = sheet.cell(row=1, column=3).value
            print (f"Содержимое ячейки в строке {1} и столбце {3}: {cell_value}")
            name = 'Дата создания'
            assert name in cell_value, f"Название столбца: {name} присутствует в файле"


def test_csv():
    with zipfile.ZipFile(ZIP_DIR, 'r') as zip_file:
        with zip_file.open("file.csv") as csv_file:
            content = csv_file.read().decode(
                'utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            second_row = csvreader[1]  # получаем вторую строку
            result_list = second_row[0].split(';')
            Postback = "OU001"
            IDPostback = 'Иванова'

            # if len(result_list) >= 3:
            #     assert result_list[
            #                2] == Postback, f"Название постбэка: {Postback} присутствует в таблице {Postback}"
            # else:
            #     print("Список result_list содержит меньше 3 элементов")

            assert result_list[2] == Postback, f"Название постбэка: {Postback
            } присутствует в таблице {result_list}"  # проверка значения элемента в первом столбце второй строки
            assert result_list[3] == IDPostback, f"ID по: {IDPostback
            } присутствует в таблице {IDPostback}"  # проверка значения элемента во втором столбце второй строки
